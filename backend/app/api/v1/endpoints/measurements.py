from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.db.session import get_db_session
from app.models.measurement import Measurement
from app.schemas.measurement import MeasurementAlertOut, MeasurementListOut, MeasurementOut, MeasurementStatsOut
from app.services.measurements import (
    aggregate_measurements,
    list_measurements,
    list_measurements_for_export,
    list_out_of_range_alerts,
)

router = APIRouter()
REPORT_TZ = ZoneInfo("Europe/Moscow")
REPORT_TZ_LABEL = "МСК (UTC+3)"


def _normalize_query_datetime(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _parse_measurement_period(
    logger_id: uuid.UUID | None,
    from_: datetime | None,
    to: datetime | None,
) -> tuple[uuid.UUID | None, datetime | None, datetime | None]:
    captured_from = _normalize_query_datetime(from_) if from_ is not None else None
    captured_to = _normalize_query_datetime(to) if to is not None else None
    if captured_from is not None and captured_to is not None and captured_from > captured_to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Parameter 'from' must be less than or equal to 'to'.",
        )
    return logger_id, captured_from, captured_to


def _media_url_or_path(public_base: str, image_path: str | None) -> str:
    if not image_path:
        return ""
    rel = f"/media/{image_path.lstrip('/')}"
    base = public_base.rstrip("/")
    if base:
        return f"{base}{rel}"
    return rel


def _fmt_optional_bool(v: bool | None) -> str:
    if v is None:
        return ""
    return "true" if v else "false"


def _export_csv_row(m: Measurement, logger_name: str, public_base: str) -> list[str]:
    val = "" if m.value is None else str(m.value)
    return [
        m.captured_at.isoformat(),
        str(m.logger_id),
        logger_name,
        val,
        m.unit,
        "true" if m.ok else "false",
        _fmt_optional_bool(m.out_of_range),
        _media_url_or_path(public_base, m.image_path),
        m.error or "",
    ]


def _format_value(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.3f}"


def _format_int(value: int | None) -> str:
    return str(int(value or 0))


def _file_ts() -> str:
    return datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")


def _format_dt_ru(dt: datetime | None) -> str:
    if dt is None:
        return "—"
    return dt.astimezone(REPORT_TZ).strftime("%d.%m.%Y %H:%M:%S")


def _get_pdf_font_name() -> str:
    """Возвращает имя шрифта с поддержкой кириллицы; fallback на Helvetica."""
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
        Path("/Library/Fonts/Arial.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
    ]
    for p in candidates:
        if not p.exists():
            continue
        font_name = f"ui_unicode_{p.stem.lower()}"
        try:
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, str(p)))
            return font_name
        except Exception:
            continue
    return "Helvetica"


@router.get(
    "/export.csv",
    summary="Экспорт измерений в CSV",
    description="Колонки: captured_at, logger_id, logger_name, value, unit, ok, out_of_range, image_url_or_path, error. "
    "Фильтры как у списка. APP_PUBLIC_BASE_URL — абсолютные ссылки на кадры; иначе относительный путь /media/….",
    response_class=StreamingResponse,
)
async def export_measurements_csv(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    max_rows: int = Query(default=100_000, ge=1, le=500_000, description="Максимум строк в выгрузке."),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    rows = await list_measurements_for_export(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        max_rows=max_rows,
    )
    public_base = settings.public_base_url.strip()

    header = [
        "captured_at",
        "logger_id",
        "logger_name",
        "value",
        "unit",
        "ok",
        "out_of_range",
        "image_url_or_path",
        "error",
    ]

    def generate_chunks():
        buf = io.StringIO()
        writer = csv.writer(buf)
        yield "\ufeff".encode("utf-8")
        writer.writerow(header)
        yield buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate(0)
        for m, logger_name in rows:
            writer.writerow(_export_csv_row(m, logger_name, public_base))
            yield buf.getvalue().encode("utf-8")
            buf.seek(0)
            buf.truncate(0)

    headers = {"Content-Disposition": 'attachment; filename="measurements_export.csv"'}
    return StreamingResponse(
        generate_chunks(),
        media_type="text/csv; charset=utf-8",
        headers=headers,
    )


@router.get(
    "/export.xlsx",
    summary="Экспорт измерений в Excel (XLSX)",
    description="Листы: summary, measurements, alerts. Фильтры как у списка.",
    response_class=StreamingResponse,
)
async def export_measurements_xlsx(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    max_rows: int = Query(default=100_000, ge=1, le=500_000, description="Максимум строк в выгрузке."),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    rows = await list_measurements_for_export(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        max_rows=max_rows,
    )
    stats = await aggregate_measurements(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
    )
    alerts = await list_out_of_range_alerts(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        limit=1000,
    )
    public_base = settings.public_base_url.strip()

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"

    ws_summary.append(["Параметр", "Значение"])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    ws_summary.append(["Часовой пояс отчёта", REPORT_TZ_LABEL])
    ws_summary.append(["Период от", _format_dt_ru(captured_from)])
    ws_summary.append(["Период до", _format_dt_ru(captured_to)])
    ws_summary.append(["Логер", str(lid) if lid else "Все"])
    ws_summary.append(["Всего записей", _format_int(stats.count)])
    ws_summary.append(["С числом", _format_int(stats.value_count)])
    ws_summary.append(["Минимум", _format_value(stats.value_min)])
    ws_summary.append(["Максимум", _format_value(stats.value_max)])
    ws_summary.append(["Среднее", _format_value(stats.value_avg)])
    ws_summary.append(["Ошибки распознавания", _format_int(stats.recognition_fail_count)])
    ws_summary.append(["Вне диапазона", _format_int(stats.out_of_range_count)])
    ws_summary.append(["Предупреждения CV", _format_int(stats.cv_warnings_count)])
    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 32

    ws_measurements = wb.create_sheet("measurements")
    ws_measurements.append(
        [
            "captured_at",
            "logger_id",
            "logger_name",
            "value",
            "unit",
            "ok",
            "out_of_range",
            "image_url_or_path",
            "error",
        ]
    )
    for cell in ws_measurements[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    for m, logger_name in rows:
        row = _export_csv_row(m, logger_name, public_base)
        row[0] = _format_dt_ru(m.captured_at)
        ws_measurements.append(row)
    ws_measurements.auto_filter.ref = f"A1:I{max(1, ws_measurements.max_row)}"
    ws_measurements.freeze_panes = "A2"
    for key, width in {"A": 24, "B": 38, "C": 24, "D": 12, "E": 10, "F": 8, "G": 12, "H": 44, "I": 36}.items():
        ws_measurements.column_dimensions[key].width = width

    ws_alerts = wb.create_sheet("alerts")
    ws_alerts.append(["captured_at", "logger_id", "logger_name", "value", "unit", "error", "image_url_or_path"])
    for cell in ws_alerts[1]:
        cell.font = Font(bold=True)
    for m, logger_name in alerts:
        ws_alerts.append(
            [
                _format_dt_ru(m.captured_at),
                str(m.logger_id),
                logger_name,
                "" if m.value is None else str(m.value),
                m.unit,
                m.error or "",
                _media_url_or_path(public_base, m.image_path),
            ]
        )
    ws_alerts.auto_filter.ref = f"A1:G{max(1, ws_alerts.max_row)}"
    ws_alerts.freeze_panes = "A2"
    for key, width in {"A": 24, "B": 38, "C": 24, "D": 12, "E": 10, "F": 36, "G": 44}.items():
        ws_alerts.column_dimensions[key].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="measurements_report_{_file_ts()}.xlsx"'}
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get(
    "/export.pdf",
    summary="Экспорт сводного отчёта в PDF",
    description="Краткая сводка за период + список последних аномалий.",
    response_class=StreamingResponse,
)
async def export_measurements_pdf(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    stats = await aggregate_measurements(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
    )
    alerts = await list_out_of_range_alerts(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        limit=20,
    )

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    font_name = _get_pdf_font_name()
    _, height = A4
    y = height - 20 * mm

    def line(text: str, size: int = 10, lead: float = 6.5) -> None:
        nonlocal y
        if y < 20 * mm:
            pdf.showPage()
            y = height - 20 * mm
        pdf.setFont(font_name, size)
        pdf.drawString(15 * mm, y, text)
        y -= lead * mm

    pdf.setTitle("Gauge Reader Report")
    line("Отчёт по измерениям", size=16, lead=8)
    line(f"Сформирован: {_format_dt_ru(datetime.now(tz=timezone.utc))} ({REPORT_TZ_LABEL})")
    line(f"Период: {_format_dt_ru(captured_from)}  ..  {_format_dt_ru(captured_to)}")
    line(f"Логер: {lid if lid else 'Все'}")
    line("")
    line("Сводка", size=13, lead=7)
    line(f"Всего записей: {_format_int(stats.count)}")
    line(f"С числовым значением: {_format_int(stats.value_count)}")
    line(f"Мин / Макс / Среднее: {_format_value(stats.value_min)} / {_format_value(stats.value_max)} / {_format_value(stats.value_avg)}")
    line(f"Ошибки распознавания: {_format_int(stats.recognition_fail_count)}")
    line(f"Вне диапазона: {_format_int(stats.out_of_range_count)}")
    line(f"Предупреждения CV: {_format_int(stats.cv_warnings_count)}")
    line("")
    line("Последние аномалии (до 20)", size=13, lead=7)
    if not alerts:
        line("Аномалий не найдено.")
    else:
        for m, logger_name in alerts:
            val = "—" if m.value is None else f"{m.value:.3f} {m.unit}"
            err = f" | {m.error}" if m.error else ""
            line(f"{_format_dt_ru(m.captured_at)} | {logger_name} | {val}{err}", size=9, lead=5.5)

    pdf.showPage()
    pdf.save()
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="measurements_report_{_file_ts()}.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


@router.get(
    "/alerts",
    response_model=list[MeasurementAlertOut],
    summary="Список алертов по аномалиям",
    description=(
        "In-app алерты по подтверждённым аномалиям: out_of_range=true, ok=true, "
        "без критичных предупреждений CV (новые сверху)."
    ),
)
async def api_measurement_alerts(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    limit: int = Query(default=50, ge=1, le=500),
    session: AsyncSession = Depends(get_db_session),
) -> list[MeasurementAlertOut]:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    rows = await list_out_of_range_alerts(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        limit=limit,
    )
    return [
        MeasurementAlertOut(
            measurement_id=m.id,
            logger_id=m.logger_id,
            logger_name=logger_name,
            captured_at=m.captured_at,
            value=m.value,
            unit=m.unit,
            error=m.error,
            image_path=m.image_path,
        )
        for m, logger_name in rows
    ]


@router.get(
    "/stats",
    response_model=MeasurementStatsOut,
    summary="Сводка по измерениям за период",
    description=(
        "Агрегаты за интервал [from, to] по логеру или по всем. "
        "Min/max/avg считаются по очищенным значениям: ok=true, без критичных cv_warnings; "
        "для analog дополнительно отсечение по шкале calibration min/max. "
        "Остальные счётчики считаются по сырым данным периода."
    ),
)
async def api_measurement_stats(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    session: AsyncSession = Depends(get_db_session),
) -> MeasurementStatsOut:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    row = await aggregate_measurements(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
    )
    return MeasurementStatsOut(
        period_from=captured_from,
        period_to=captured_to,
        logger_id=lid,
        count=row.count,
        value_count=row.value_count,
        value_min=row.value_min,
        value_max=row.value_max,
        value_avg=row.value_avg,
        recognition_fail_count=row.recognition_fail_count,
        out_of_range_count=row.out_of_range_count,
        cv_warnings_count=row.cv_warnings_count,
    )


@router.get("/", response_model=MeasurementListOut)
async def api_list_measurements(
    logger_id: uuid.UUID | None = Query(default=None),
    from_: datetime | None = Query(default=None, alias="from", description="Начало периода (captured_at), включительно (UTC)."),
    to: datetime | None = Query(default=None, description="Конец периода (captured_at), включительно (UTC)."),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    session: AsyncSession = Depends(get_db_session),
) -> MeasurementListOut:
    lid, captured_from, captured_to = _parse_measurement_period(logger_id, from_, to)
    items, total = await list_measurements(
        session,
        logger_id=lid,
        captured_from=captured_from,
        captured_to=captured_to,
        offset=offset,
        limit=limit,
    )
    return MeasurementListOut(items=[MeasurementOut.model_validate(m) for m in items], total=total)

